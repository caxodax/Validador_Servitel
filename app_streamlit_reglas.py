# app_streamlit_reglas.py
import io, os, re, yaml, difflib, datetime
import pandas as pd
import streamlit as st
from pathlib import Path
from validator import apply_non_destructive_normalization, apply_auto_fixes, validate_dataframe
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


MAESTRO_SHEET = "MAESTRO"

# ----- Posiciones (0-based) -----
UPPERCASE_N = 4            # columnas 0..3 -> MAYÚSCULAS
TIPO_DOC_POS = 5           # col 6 (1-based) -> *TIPO DOCUMENTO*
DOC_NUM_POS  = 6           # col 7 (1-based) -> *N. DOCUMENTO*
EMAIL_POS    = 7           # col 8 (1-based) -> *E-MAIL*
# FECHAS por POSICIÓN 0-based (obligatorias, dd/mm/yyyy)
DATE_POSITIONS = [12, 13, 14, 15]

# ----- Columnas opcionales -----
OPTIONAL_COLUMNS = {
    "ZONA",
    "TIPO CALLE",
    "TIPO DE CONJUNTO",
    "TIPO VIVIENDA",
    "RUTA",
    "ORDEN RUTA",
    "CALLE PRINCIPAL",
    "NUMERO DE CASA",
    "CALLE SECUNDARIA",
    "LATITUD",
    "LONGITUD",
}

SEVERITY_MAP_ES = {
    "error": "Error",
    "warn": "Aviso",
    "info": "Información"
}
def dedupe_errors_df(err_df: pd.DataFrame) -> pd.DataFrame:
    """
    Elimina incidencias duplicadas por celda.
    Clave principal: (fila, col_idx) si col_idx existe.
    Fallback: (fila, columna).

    Preferencia:
      - Regla más específica gana (telefono_vacio, ip_invalida, etc.)
      - Si empata, gana la severidad más alta (error > warn > info)
      - Si empata, gana el detalle más largo
    """
    if err_df is None or err_df.empty:
        return pd.DataFrame(columns=["fila","columna","col_idx","regla","detalle","severity"])

    df = err_df.copy()

    # Normalizar columnas esperadas
    for col in ["fila","columna","col_idx","regla","detalle","severity"]:
        if col not in df.columns:
            df[col] = None

    # normalizaciones mínimas
    df["fila"] = pd.to_numeric(df["fila"], errors="coerce").fillna(-1).astype(int)
    df["columna"] = df["columna"].astype(str)
    df["regla"] = df["regla"].astype(str)
    df["detalle"] = df["detalle"].astype(str)
    df["severity"] = df["severity"].astype(str).str.lower().replace({"warning":"warn"})

    # Prioridad por regla (más específico > genérico)
    rule_priority = {
        "telefono_vacio": 100,
        "telefono_invalido": 95,
        "email_vacio": 90,
        "email_sin_arroba": 85,
        "email_invalido": 80,
        "fecha_obligatoria": 80,
        "mac_vacia": 80,
        "mac_invalida": 75,
        "ip_invalida": 70,
        "vlan_invalida": 65,
        "ont_vacio": 80,
        "ont_longitud_invalida": 75,
        "ont_formato_invalido": 70,
        "campo_obligatorio": 10,     # genérica (pierde contra casi todo)
    }

    sev_priority = {"error": 3, "warn": 2, "info": 1}

    # Key por celda: (fila, col_idx) si col_idx es válido, si no: (fila, columna)
    has_col_idx = "col_idx" in df.columns
    if has_col_idx:
        df["_col_idx_valid"] = pd.to_numeric(df["col_idx"], errors="coerce")
        df["_key"] = df.apply(
            lambda r: (int(r["fila"]), int(r["_col_idx_valid"])) if pd.notna(r["_col_idx_valid"]) else (int(r["fila"]), str(r["columna"])),
            axis=1
        )
    else:
        df["_key"] = df.apply(lambda r: (int(r["fila"]), str(r["columna"])), axis=1)

    df["_rp"] = df["regla"].map(lambda x: rule_priority.get(x, 50))
    df["_sp"] = df["severity"].map(lambda x: sev_priority.get(x, 2))
    df["_dl"] = df["detalle"].map(lambda x: len(str(x)))

    # Orden: por key, luego mejor regla, luego severidad, luego detalle
    df = df.sort_values(by=["_key","_rp","_sp","_dl"], ascending=[True, False, False, False])

    # Mantener 1 por celda
    df = df.drop_duplicates(subset=["_key"], keep="first")

    # limpiar auxiliares
    df = df.drop(columns=[c for c in ["_key","_rp","_sp","_dl","_col_idx_valid"] if c in df.columns], errors="ignore")
    return df

# ----- Columnas deben ser entero -----
INT_COLUMNS = [
  #  "* CODIGO CONTRATO",
  #  "* N. DOCUMENTO",
    "* DURACION (MESES)",
    "* DIA COBRO",
    "* DIA CORTE",
    "PENALIDAD",
    "ORDEN RUTA",
    "NUMERO DE CASA",
    "NUMERO DE CALLE",
    "PRECIO DE TV ANALOGA",
    "PRECIO TV ADICIONALES",
    "DEUDA TV",
    "PRECIO DE INTERNET",
    "DEUDA INTERNET",
    "PRECIO DE CATV",
    "DEUDA CATV",
    "PRECIO DE IPTV",
    "DEUDA IPTV",
    "PRECIO DE TELEFONIA",
    "DEUDA TELEFONIA",
    "PRECIO DE CAJA DIGITAL",
    "DEUDA CAJA DIGITAL",
    "VALOR"
]


# ----- Dominios de e-mail permitidos -----
ALLOWED_DOMAINS = [
    "gmail.com",
    "hotmail.com", "hotmail.es",
    "outlook.com", "outlook.es",
    "yahoo.com",
]
ALLOWED_DOMAINS_STRIPPED = {d.replace(".", ""): d for d in ALLOWED_DOMAINS}
COMMON_DOMAIN_TYPO_MAP = {
    "gmai.com": "gmail.com",
    "gmaii.com": "gmail.com",
    "gmal.com": "gmail.com",
    "gmail.con": "gmail.com",
    "gmail,com": "gmail.com",
    "hotmail.con": "hotmail.com",
    "hotmai.com": "hotmail.com",
    "hotmal.com": "hotmail.com",
    "hotmail,es": "hotmail.es",
    "outlok.com":  "outlook.com",
    "outlok.es":   "outlook.es",
    "outlook,es":  "outlook.es",
    "yahho.com":   "yahoo.com",
    "yaho.com":    "yahoo.com",
}

# === FAVICON PERSONALIZADO ===
import base64
from pathlib import Path

BASE_DIR = Path(__file__).parent
assets_dir = BASE_DIR / "assets"
favicon_path = assets_dir / "icono_servitel.ico"

favicon_bytes = None
if favicon_path.exists():
    favicon_bytes = open(favicon_path, "rb").read()

# === CONFIGURACIÓN DE PÁGINA ===
st.set_page_config(
    page_title="Validador - Servitel",
    page_icon=favicon_bytes,   # ← tu .ico real convertido a bytes
    layout="wide",
    initial_sidebar_state="collapsed"
)



BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def find_logo(*relative_paths):
    """
    Devuelve la primera ruta existente dentro de BASE_DIR de la lista dada.
    Ejemplo de uso:
      find_logo("assets/logo_servitel.jpg", "assets/logo_servitel.png")
    """
    for rel in relative_paths:
        path = os.path.join(BASE_DIR, rel)
        if os.path.exists(path):
            return path
    return None

# --- Encabezado con logo Servitel centrado y título ---
BASE_DIR = Path(__file__).parent
assets_dir = BASE_DIR / "assets"

col_left, col_center, col_right = st.columns([1, 2, 1])

with col_center:
    servitel_logo = None
    if assets_dir.exists():
        candidates = list(assets_dir.glob("logo_servitel*"))
        if candidates:
            servitel_logo = candidates[0]

    if servitel_logo is not None:
        st.markdown(
            f"""
            <div style="display:flex; justify-content:center;">
                <img src="data:image/png;base64,{base64.b64encode(open(servitel_logo, "rb").read()).decode()}"
                    style="width:240px;"/>
            </div>
            """,
            unsafe_allow_html=True
        )

    # Título centrado debajo del logo
    st.markdown(
        "<h1 style='text-align:center; margin-bottom: 0;'>Sistema interno de validación de contratos</h1>",
        unsafe_allow_html=True,
    )
    

# =================== LECTURA INTELIGENTE DE ENCABEZADOS ===================
HEADER_ROW_INDEX = 1  # segunda fila (0-based)

def _norm_cell_token(x: str) -> str:
    s = "" if x is None else str(x)
    s = s.replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    # quita asteriscos iniciales y el espacio siguiente
    s = re.sub(r"^\*+\s*", "", s)
    return s.upper()

def _trim_trailing_empty_rows(df: pd.DataFrame, anchor_col_name: str) -> pd.DataFrame:
    """
    Recorta el DF hasta la última fila donde anchor_col_name tiene valor real.
    Evita miles de errores falsos por filas vacías formateadas en Excel.
    """
    if anchor_col_name not in df.columns:
        return df

    ser = df[anchor_col_name].astype(str).map(lambda v: v.strip())
    # considera vacíos: "", "nan", "none", "null", "<na>"
    is_real = ~ser.str.lower().isin({"", "nan", "none", "null", "<na>"})
    if not is_real.any():
        return df

    last_idx = is_real[is_real].index.max()
    out = df.loc[:last_idx].copy()
    out.reset_index(drop=True, inplace=True)
    return out

def smart_read_maestro(file) -> pd.DataFrame:
    """
    Lector inteligente:
    - Encuentra la fila header buscando 'CODIGO CONTRATO' (normalizado).
    - Si no la encuentra, cae a la heurística anterior.
    - Limpia columnas 'Unnamed'
    - Recorta filas vacías al final usando '* CODIGO CONTRATO' (o equivalente normalizado) como ancla.
    """
    df0 = pd.read_excel(file, sheet_name=MAESTRO_SHEET, header=None, engine="openpyxl")

    target = "CODIGO CONTRATO"
    header_row = None

    max_scan = min(20, len(df0))  # un poco más de margen
    for i in range(max_scan):
        row_norm = df0.iloc[i].map(_norm_cell_token)
        if (row_norm == target).any():
            header_row = i
            break

    # fallback a heurística anterior
    if header_row is None:
        first_row = df0.iloc[0].astype(str)
        many_unnamed = (first_row.str.startswith("Unnamed")).mean() > 0.5
        has_title = first_row.str.contains("INFORMACION", case=False, na=False).any()
        header_row = HEADER_ROW_INDEX if (many_unnamed or has_title) else 0

        # si en la fila 2 aparece CODIGO CONTRATO, úsala
        if len(df0) > 1:
            row1_norm = df0.iloc[1].map(_norm_cell_token)
            if (row1_norm == target).any():
                header_row = 1

    header = df0.iloc[header_row].astype(str).str.strip().tolist()
    df = df0.iloc[header_row + 1:].copy()
    df.columns = header
    df.reset_index(drop=True, inplace=True)

    # limpiar 'Unnamed'
    df.columns = [c if not str(c).lower().startswith("unnamed") else "" for c in df.columns]

    # encontrar la columna ancla (la que normaliza a 'CODIGO CONTRATO')
    anchor = None
    for c in df.columns:
        if _norm_cell_token(c) == target:
            anchor = c
            break

    if anchor:
        df = _trim_trailing_empty_rows(df, anchor)

    return df

# =================== CARGA DE REGLAS (estructura/fechas) ===================
@st.cache_data
def load_rules():
    with open("rules/schema_rules.yaml", "r", encoding="utf-8") as f:
        return yaml.safe_load(f)
rules_pkg = load_rules()
COLUMN_ORDER = rules_pkg.get("column_order", None)

# ---------- Normalización de nombres de columnas para comparación ----------
def normalize_colname_for_match(name: str) -> str:
    """
    Normaliza nombres de columnas para comparación lógica:
    - Convierte a str.
    - Reemplaza espacios raros.
    - Colapsa espacios múltiples.
    - Quita asteriscos iniciales (*OBSERVACIONES -> OBSERVACIONES).
    - Pasa a MAYÚSCULAS.
    """
    s = str(name)
    s = s.replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    # Quitar asterisco(s) inicial(es) y espacio que le siga, si aplica
    s = re.sub(r"^\*+\s*", "", s)
    return s.upper()


def compare_columns_with_details(expected, received):
    """
    Compara listas de columnas usando la versión normalizada para igualdad,
    y devuelve:
      - ok: bool -> True si coinciden (tras normalizar).
      - diff_df: DataFrame con las posiciones donde NO coinciden.
    """
    norm_expected = [normalize_colname_for_match(c) for c in expected]
    norm_received = [normalize_colname_for_match(c) for c in received]

    ok = norm_expected == norm_received

    diffs = []
    max_len = max(len(expected), len(received))
    for i in range(max_len):
        exp_raw  = expected[i] if i < len(expected) else ""
        rec_raw  = received[i] if i < len(received) else ""
        exp_norm = norm_expected[i] if i < len(norm_expected) else ""
        rec_norm = norm_received[i] if i < len(norm_received) else ""

        if exp_norm != rec_norm:
            diffs.append({
                "posición": i,
                "esperado": exp_raw,
                "recibido": rec_raw,
                "esperado_norm": exp_norm,
                "recibido_norm": rec_norm,
            })

    diff_df = pd.DataFrame(diffs)
    return ok, diff_df


# ---------- Helpers encabezados (solo visual) ----------
def is_unnamed(col_name) -> bool:
    return str(col_name).lower().startswith("unnamed")

def run_full_pipeline(df_in: pd.DataFrame, enable_autocorrect: bool):
    """
    Aplica el mismo pipeline usado al cargar:
    - uppercase primeras columnas
    - inferir tipo documento
    - normalizar texto (acentos)
    - solo dígitos en N. DOCUMENTO
    - limpiar MAC
    - normalizar VLAN
    - autocorrect email (si está habilitado)
    - convertir int columnas
    Retorna (df_out, tipo_errs_all, corrections_df, info_email_errors)
    """
    df = df_in.copy()

    df, _ = force_uppercase_first_cols(df, n=UPPERCASE_N)

    df, tipo_errs_all = infer_tipo_documento_from_docnum(df, TIPO_DOC_POS, DOC_NUM_POS)

    df, _ = normalize_text_apostrophe(df)

    df = digits_only_column(df, DOC_NUM_POS)

    if "MAC" in df.columns:
        df["MAC"] = df["MAC"].map(clean_mac)

    df = normalize_vlan(df)

    corrections_df = pd.DataFrame()
    info_email_errors = pd.DataFrame(columns=["fila","columna","col_idx","regla","detalle","severity"])
    if enable_autocorrect:
        df, corrections_df, info_email_errors = apply_email_autocorrect(df, EMAIL_POS, True)

    df = convert_int_columns(df, INT_COLUMNS)

    return df, tipo_errs_all, corrections_df, info_email_errors

def build_column_config(df: pd.DataFrame):
    cfg = {}
    for c in df.columns:
        label = "" if is_unnamed(c) else str(c)
        cfg[c] = st.column_config.Column(label=label)
    return cfg

# ------- Resalta el error de columnas ---------------
def highlight_column_list(columns, diff_df):
    """
    Devuelve una lista con:
    - 🔴 emoji de alerta para columnas en diff_df
    - Texto en negrita
    - Fondo rojo suave para columnas incorrectas
    - Formato: LETRA_EXCEL: nombre de columna (ej: AF: ESTADO)
    """
    bad_positions = set(diff_df["posición"].tolist())
    styled = []

    for i, col in enumerate(columns):
        letra = idx_to_excel_col(i)  # convierte 0,1,2... -> A,B,C...
        label = f"{letra}: {col}"

        if i in bad_positions:
            styled.append(
                f"<span style='background-color:#ffcccc; padding:3px 6px; border-radius:4px;'>"
                f"🔴 <b>{label}</b>"
                f"</span>"
            )
        else:
            styled.append(label)

    return styled


# ---------- Convertir columnas decimal a entero ----------
def convert_int_columns(df: pd.DataFrame, int_cols: list[str]) -> pd.DataFrame:
    df2 = df.copy()
    for col in int_cols:
        if col in df2.columns:
            df2[col] = (
                pd.to_numeric(df2[col], errors="coerce")
                .fillna("")  # si está vacío, lo dejamos vacío
            )
            # quitar .0 si es entero
            df2[col] = df2[col].apply(lambda x: int(x) if isinstance(x, float) and x.is_integer() else x)
    return df2

# ---------- Columnas duplicadas: helpers ----------
def make_unique_labels(labels):
    seen = {}
    out = []
    for name in labels:
        name = "" if name is None else str(name)
        if name not in seen:
            seen[name] = 1
            out.append(name)
        else:
            seen[name] += 1
            out.append(f"{name} ({seen[name]})")
    return out

def ui_df_with_unique_columns(df: pd.DataFrame):
    """Devuelve una copia con nombres únicos SOLO para UI (misma forma y orden)."""
    dfd = df.copy()
    dfd.columns = make_unique_labels(df.columns)
    return dfd

# ---------- Helpers fechas ----------
def date_columns_from_rules(rules_pkg):
    return [r["campo"] for r in rules_pkg.get("rules", []) if str(r.get("tipo")) == "date"]

def ensure_datetime_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Convierte columnas de fecha a datetime (para escribir sin hora)."""
    out = df.copy()
    for col in date_columns_from_rules(rules_pkg):
        if col in out.columns:
            out[col] = pd.to_datetime(out[col], errors="coerce")
    for pos in DATE_POSITIONS:
        if pos < len(out.columns):
            c = out.columns[pos]
            out[c] = pd.to_datetime(out[c], errors="coerce")
    return out

def format_dates_for_display(df: pd.DataFrame, rules_pkg):
    """
    Solo para mostrar en la UI: dd/mm/yyyy.
    Soporta columnas duplicadas usando SIEMPRE posición (iloc)
    para evitar el error 'cannot assemble with duplicate keys'.
    """
    out = df.copy()

    # 1) Columnas de fecha definidas en las rules (por NOMBRE, pero aplicando por ÍNDICE)
    date_cols = date_columns_from_rules(rules_pkg)
    for col_name in date_cols:
        if col_name in out.columns:
            # puede haber columnas duplicadas con el mismo nombre
            idxs = [j for j, name in enumerate(out.columns) if name == col_name]
            for j in idxs:
                serie = pd.to_datetime(out.iloc[:, j], errors="coerce")
                out.iloc[:, j] = serie.dt.strftime("%d/%m/%Y")

    # 2) Columnas de fecha definidas por POSICIÓN (DATE_POSITIONS)
    for pos in DATE_POSITIONS:
        if pos < out.shape[1]:
            serie = pd.to_datetime(out.iloc[:, pos], errors="coerce")
            out.iloc[:, pos] = serie.dt.strftime("%d/%m/%Y")

    return out

# ---------- Transformaciones ----------
def force_uppercase_first_cols(df: pd.DataFrame, n: int) -> tuple[pd.DataFrame, int]:
    """
    Pone en MAYÚSCULAS las primeras n columnas y devuelve:
      - df2: DataFrame modificado
      - changed_count: número de celdas que cambiaron realmente
    """
    df2 = df.copy()
    changed_count = 0

    cols = list(df2.columns)[:max(0, n)]
    for c in cols:
        col = df2[c]

        # Nueva columna en mayúsculas (manteniendo NaN)
        new_col = col.map(lambda v: v if pd.isna(v) else str(v).upper())

        # Contar diferencias reales (ignorando NaN)
        try:
            mask = (~col.isna()) & (col.astype(str) != new_col.astype(str))
            changed_count += int(mask.sum())
        except Exception:
            # Si algo raro pasa con los tipos, asumimos sin conteo fino
            pass

        df2[c] = new_col

    return df2, changed_count

def digits_only_column(df: pd.DataFrame, pos: int) -> pd.DataFrame:
    if pos >= len(df.columns):
        return df
    c = df.columns[pos]
    df2 = df.copy()
    df2[c] = df2[c].map(lambda v: v if pd.isna(v) else re.sub(r"[^0-9]", "", str(v)))
    return df2

def clean_mac(value: str) -> str:
    """
    Limpia y formatea la MAC:
    - Mantiene solo caracteres hexadecimales
    - Convierte a MAYÚSCULAS
    - Inserta ':' cada 2 caracteres → XX:XX:XX:XX:XX:XX
    """
    if pd.isna(value):
        return ""

    # Solo hex en mayúsculas
    s = re.sub(r'[^0-9A-Fa-f]', '', str(value)).upper()

    if s == "":
        return ""

    # Partir en grupos de 2 caracteres
    parts = [s[i:i+2] for i in range(0, len(s), 2)]
    return ":".join(parts)



def normalize_vlan(df: pd.DataFrame) -> pd.DataFrame:
    if "VLAN" not in df.columns:
        return df
    df2 = df.copy()
    c = "VLAN"
    for i, v in df2[c].items():
        if pd.isna(v) or str(v).strip() == "":
            continue
        s = str(v).strip()
        # convertir 1003.0 -> 1003
        if re.fullmatch(r"\d+\.0+", s):
            df2.at[i, c] = s.split(".")[0]
        # eliminar caracteres NO numéricos
        df2.at[i, c] = re.sub(r"[^\d]", "", str(df2.at[i, c]))
    return df2

# ---------- Normalización de texto: acentos/ñ -> con apóstrofe pegado ----------
def normalize_text_apostrophe(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """
    Reemplaza acentos y ñ por su versión con apóstrofe pegado en TODAS las columnas de texto,
    excepto en columnas sensibles (E-MAIL/EMAIL, IP, MAC).

    Devuelve:
      - df2: DataFrame modificado
      - changed_count: número de celdas que cambiaron realmente
    """
    df2 = df.copy()
    changed_count = 0

    # columnas a excluir por nombre (en mayúsculas para comparar)
    excluded_cols = {"* E-MAIL", "E-MAIL", "EMAIL", "IP", "MAC"}

    repl = {
        "á": "a'", "é": "e'", "í": "i'", "ó": "o'", "ú": "u'",
        "Á": "A'", "É": "E'", "Í": "I'", "Ó": "O'", "Ú": "U'",
        "ñ": "n", "Ñ": "N",
    }

    def _fix_text(val):
        if pd.isna(val):
            return val
        s = str(val)
        for k, v in repl.items():
            s = s.replace(k, v)
        return s

    # por posición (para no chocar con nombres duplicados)
    for j in range(df2.shape[1]):
        col_name_upper = str(df2.columns[j]).strip().upper()
        if col_name_upper in excluded_cols:
            continue

        col = df2.iloc[:, j]
        if pd.api.types.is_object_dtype(col) or pd.api.types.is_string_dtype(col):
            new_col = col.map(_fix_text)

            # Contar diferencias reales (ignorando NaN)
            try:
                mask = (~col.isna()) & (col.astype(str) != new_col.astype(str))
                changed_count += int(mask.sum())
            except Exception:
                pass

            df2.iloc[:, j] = new_col

    return df2, changed_count

DOC_PREFIX_MAP = {"J": "JURIDICO", "V": "CEDULA", "G": "GOBIERNO", "E": "EXTRANJERO"}

def infer_tipo_documento_from_docnum(df: pd.DataFrame, tipo_pos: int, doc_pos: int):
    """
    Rellena la columna *TIPO DOCUMENTO* (posición tipo_pos) en función del contenido
    de *N. DOCUMENTO* (posición doc_pos), usando el prefijo de la cédula/RIF.

    Trabaja SIEMPRE por posición (.iat) para evitar problemas con columnas duplicadas.
    Devuelve:
      - df2: DataFrame con los tipos de documento inferidos
      - errs_df: DataFrame de errores (fila, columna, col_idx, regla, detalle, severity)
    """
    df2 = df.copy()
    errs = []

    # Seguridad: si alguna posición se sale del rango, devolvemos sin tocar nada
    if tipo_pos >= len(df2.columns) or doc_pos >= len(df2.columns):
        return df2, pd.DataFrame(
            columns=["fila", "columna", "col_idx", "regla", "detalle", "severity"]
        )

    # Nombres "lógicos" solo para reportar en errores
    tipo_col_name = df2.columns[tipo_pos]
    doc_col_name = df2.columns[doc_pos]

    # Recorremos por índice entero (0-based)
    for i in range(len(df2)):
        val = df2.iat[i, doc_pos]  # valor crudo de N. DOCUMENTO en esa fila

        # Normalizamos a string
        s = "" if pd.isna(val) else str(val).strip()

        if s == "":
            # Documento vacío -> no podemos inferir tipo
            errs.append(
                (
                    i,
                    tipo_col_name,
                    tipo_pos,
                    "tipo_documento_desconocido",
                    f"{doc_col_name} vacío",
                    "error",
                )
            )
            continue

        # Buscamos la primera letra en el documento
        m = re.search(r"[A-Za-z]", s)
        if not m:
            # Sin letras: asumimos CÉDULA (Venezolano)
            df2.iat[i, tipo_pos] = "CEDULA"
            continue

        pref = m.group(0).upper()
        mapped = DOC_PREFIX_MAP.get(pref)

        if mapped is None:
            # Prefijo no reconocido
            errs.append(
                (
                    i,
                    tipo_col_name,
                    tipo_pos,
                    "tipo_documento_desconocido",
                    f"Prefijo '{pref}' no reconocido en {doc_col_name}",
                    "error",
                )
            )
        else:
            # Prefijo reconocido -> asignamos tipo
            df2.iat[i, tipo_pos] = mapped

    if not errs:
        return df2, pd.DataFrame(
            columns=["fila", "columna", "col_idx", "regla", "detalle", "severity"]
        )

    errs_df = pd.DataFrame(
        errs, columns=["fila", "columna", "col_idx", "regla", "detalle", "severity"]
    )
    return df2, errs_df


# ---------- E-mail: normalización + autocorrección ----------
EMAIL_REGEX = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')

def _clean_email_text(s: str) -> str:
    s = s.strip()
    s = s.replace(" ", "").replace(";", ",").lower()
    s = s.replace(",",".")
    s = s.replace("..", ".")
    s = s.replace("@@", "@").replace("@.", "@")
    return s

def _insert_at_using_suffix_match(s: str) -> str | None:
    if "@" in s:  # ya trae @
        return s
    best = None
    best_ratio = 0.0
    best_pos = None
    best_domain = None
    for d in ALLOWED_DOMAINS:
        L = len(d)
        for k in range(max(1, len(s)-L-2), len(s)):
            suf = s[k:]
            ratio = difflib.SequenceMatcher(None, suf, d).ratio()
            if ratio > best_ratio:
                best_ratio, best, best_pos, best_domain = ratio, suf, k, d
    if best is not None and best_ratio >= 0.72:
        local = s[:best_pos]
        if local:
            return f"{local}@{best_domain}"
    return None

def autocorrect_email(value):
    """
    Devuelve (new_value, corrected:bool, detail:str, valid:bool).

    Esta versión evita usar pd.isna(value) para no provocar
    'ValueError: The truth value of a Series is ambiguous'
    cuando el valor no es escalar.
    """
    # 1) Normalizamos SIEMPRE a string, sin usar pd.isna
    if value is None:
        original = ""
    else:
        try:
            original = str(value)
        except Exception:
            original = ""

    s = _clean_email_text(original)

    # 2) Consideramos nulos o vacíos: "", none, nan, null, <na>
    if s == "" or s.lower() in {"none", "nan", "null", "<na>"}:
        return original, False, "", False

    # Sin '@': intento insertar por sufijo aproximado o por dominio pegado.
    if "@" not in s:
        candidate = _insert_at_using_suffix_match(s)
        if candidate:
            s = candidate
        else:
            for d in ALLOWED_DOMAINS:
                idx = s.rfind(d)
                if idx > 0:
                    local = s[:idx]
                    if local:
                        s = f"{local}@{d}"
                        break

    # Si trae más de un '@', dejo solo el primero.
    if s.count("@") > 1:
        parts = s.split("@")
        s = parts[0] + "@" + "".join(parts[1:])

    if "@" not in s:
        # No se pudo construir un correo válido
        return s, False, "", False

    local, domain = s.split("@", 1)
    if local == "":
        return s, False, "", False

    corrected = False
    detail = ""

    # Correcciones comunes directas
    if domain in COMMON_DOMAIN_TYPO_MAP:
        new_domain = COMMON_DOMAIN_TYPO_MAP[domain]
        detail = f"{domain} → {new_domain}"
        domain = new_domain
        corrected = True

    # Si ya está en la lista de permitidos
    if domain in ALLOWED_DOMAINS:
        final_email = f"{local}@{domain}"
        return final_email, corrected, detail, EMAIL_REGEX.fullmatch(final_email) is not None

    # Fuzzy contra dominios permitidos
    dom_stripped = domain.replace(".", "")
    candidates = list(ALLOWED_DOMAINS_STRIPPED.keys())
    close = difflib.get_close_matches(dom_stripped, candidates, n=1, cutoff=0.8)
    if close:
        best = ALLOWED_DOMAINS_STRIPPED[close[0]]
        detail = f"{domain} → {best}" if not detail else f"{detail}; {domain} → {best}"
        domain = best
        corrected = True
    else:
        if dom_stripped in ALLOWED_DOMAINS_STRIPPED:
            best = ALLOWED_DOMAINS_STRIPPED[dom_stripped]
            detail = f"{domain} → {best}" if not detail else f"{detail}; {domain} → {best}"
            domain = best
            corrected = True

    final_email = f"{local}@{domain}"
    valid = EMAIL_REGEX.fullmatch(final_email) is not None and (domain in ALLOWED_DOMAINS)
    return final_email, corrected, detail, valid


def apply_email_autocorrect(df: pd.DataFrame, email_pos: int, enable: bool):
    corr_rows = []
    info_errors = []
    if not enable or email_pos >= len(df.columns):
        return df, pd.DataFrame(columns=["fila","original","nuevo","detalle","timestamp"]), pd.DataFrame(columns=["fila","columna","col_idx","regla","detalle","severity"])

    c = df.columns[email_pos]
    for i in df.index:
        orig = df.at[i, c]
        new_val, corrected, detail, valid = autocorrect_email(orig)
        if corrected:
            df.at[i, c] = new_val
            corr_rows.append({
                "fila": i,
                "original": orig,
                "nuevo": new_val,
                "detalle": detail if detail else "Autocorrección",
                "timestamp": datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z",
            })
            info_errors.append((i, c, email_pos, "email_autocorregido", detail or "Corrección automática de dominio", "info"))

    corr_df = pd.DataFrame(corr_rows, columns=["fila","original","nuevo","detalle","timestamp"])
    info_df = pd.DataFrame(info_errors, columns=["fila","columna","col_idx","regla","detalle","severity"]) if info_errors else pd.DataFrame(columns=["fila","columna","col_idx","regla","detalle","severity"])

    # Guardar historial
    if not corr_df.empty:
        try:
            os.makedirs("history", exist_ok=True)
            path = os.path.join("history", "email_corrections.csv")
            append_header = not os.path.exists(path)
            corr_to_save = corr_df.copy()
            corr_to_save["fila"] = corr_to_save["fila"].astype(int)
            corr_to_save.to_csv(path, index=False, mode="a", header=append_header, encoding="utf-8")
        except Exception:
            pass

    return df, corr_df, info_df

# ---------- Validaciones ----------
IP_COL_NAME = "IP"
IP_REGEX = re.compile(r'^(?:(?:25[0-5]|2[0-4][0-9]|1?[0-9]?[0-9])\.){3}(?:25[0-5]|2[0-4][0-9]|1?[0-9]?[0-9])$')
PHONE_COL_NAME = "TELEFONOS"  # columna de teléfonos
MAC_COL_NAME = "MAC"
# Formato MAC estándar: XX:XX:XX:XX:XX:XX (hexadecimal en mayúsculas)
MAC_REGEX = re.compile(r'^([0-9A-F]{2}:){5}[0-9A-F]{2}$')


def validate_position_rules(df: pd.DataFrame) -> pd.DataFrame:
    errs = []

    # ================== EMAIL OBLIGATORIO + DOMINIO PERMITIDO ==================
    if EMAIL_POS < len(df.columns):
        c = df.columns[EMAIL_POS]  # solo para mostrar el nombre de la columna
        # Usar SIEMPRE posición para evitar problemas con nombres duplicados
        series = df.iloc[:, EMAIL_POS].astype(str).str.strip()

        for i, v in series.items():
            ok_regex = EMAIL_REGEX.fullmatch(v or "")
            ok_domain = False

            if ok_regex:
                try:
                    domain = v.split("@", 1)[1].lower()
                    ok_domain = domain in ALLOWED_DOMAINS
                except Exception:
                    ok_domain = False

            if (v == "" or v.lower() in {"none", "nan", "null"}):
                errs.append((i, c, EMAIL_POS, "email_vacio", "Correo vacío", "error"))
            elif "@" not in v:
                errs.append((i, c, EMAIL_POS, "email_sin_arroba", "Correo sin '@' o dominio", "error"))
            elif (not ok_regex) or (not ok_domain):
                errs.append((i, c, EMAIL_POS, "email_invalido", "Correo inválido o dominio no permitido", "warn"))

    # ================== FECHAS OBLIGATORIAS (por POSICIÓN) ==================
    for pos in DATE_POSITIONS:
        if pos < len(df.columns):
            c = df.columns[pos]
            # IMPORTANTE: usar iloc para evitar DataFrames por columnas duplicadas
            series = pd.to_datetime(df.iloc[:, pos], errors="coerce")
            for i, val in series.items():
                if pd.isna(val):
                    errs.append(
                        (
                            i,
                            c,
                            pos,
                            "fecha_obligatoria",
                            "Fecha vacía o inválida (dd/mm/yyyy)",
                            "error",
                        )
                    )

    # ================== IP VÁLIDA ==================
    if IP_COL_NAME in df.columns:
        c = IP_COL_NAME
        col_idx = df.columns.get_loc(c)
        for i, v in df[c].items():
            if pd.isna(v) or str(v).strip() == "":
                continue  # si quieres que sea obligatoria, cambia a error aquí
            val = str(v).strip()
            if not IP_REGEX.fullmatch(val):
                errs.append(
                    (
                        i,
                        df.columns[col_idx],
                        col_idx,
                        "ip_invalida",
                        "IP inválida: debe tener 4 segmentos 0–255 (ej. 172.18.9.10)",
                        "error",
                    )
                )

    # ================== TELEFONOS ==================
    if PHONE_COL_NAME in df.columns:
        c = PHONE_COL_NAME
        col_idx = df.columns.get_loc(c)

        def _to_digits(val) -> str:
            if pd.isna(val):
                return ""
            # Si excel lo trae numérico (float) evita "....0" o notación científica
            try:
                import numpy as np
                if isinstance(val, (int, np.integer)):
                    return str(int(val))
                if isinstance(val, (float, np.floating)):
                    if float(val).is_integer():
                        return str(int(val))
                    # si no es entero, igual sacamos dígitos del string
            except Exception:
                pass
            return re.sub(r"\D", "", str(val))

        for i, v in df[c].items():
            s_raw = "" if pd.isna(v) else str(v)
            s = s_raw.replace("−", "-").replace("–", "-")
            s = re.sub(r"\s+", " ", s).strip()

            if s == "":
                errs.append(
                    (
                        i,
                        c,
                        col_idx,
                        "telefono_vacio",
                        "Teléfono vacío o incompleto (se esperan 10 u 11 dígitos).",
                        "error",
                    )
                )
                continue

            raw_parts = s.split("/")
            parts = [p.strip() for p in raw_parts if p.strip() != ""]

            if not parts:
                errs.append(
                    (
                        i,
                        c,
                        col_idx,
                        "telefono_invalido",
                        "Teléfono inválido (use 10 u 11 dígitos, opcionalmente separados por '/').",
                        "error",
                    )
                )
                continue

            normalized_parts = []
            ok = True
            corrected = False

            for p in parts:
                digits = _to_digits(p)

                # Si viene 10 dígitos (ej. 414xxxxxxx), le agregamos 0 → 0414xxxxxxx
                if len(digits) == 10:
                    digits = "0" + digits
                    corrected = True

                # Aceptamos solo 11 al final
                if len(digits) != 11:
                    ok = False
                    break

                normalized_parts.append(digits)

            if not ok:
                errs.append(
                    (
                        i,
                        c,
                        col_idx,
                        "telefono_invalido",
                        "Cada número debe tener 10 u 11 dígitos (ej. 414xxxxxxx o 0414xxxxxxx). "
                        "Si viene con 10, el sistema agregará el 0.",
                        "error",
                    )
                )
            else:
                # Si se corrigió (o si quieres siempre normalizar), guarda normalizado
                if corrected:
                    df.at[i, c] = " / ".join(normalized_parts)


    # ================== MAC (OBLIGATORIA, FORMATO XX:XX:XX:XX:XX:XX) ==================
    if MAC_COL_NAME in df.columns:
        c = MAC_COL_NAME
        col_idx = df.columns.get_loc(c)

        for i, v in df[c].items():
            s = "" if pd.isna(v) else str(v).strip().upper()

            # No puede estar vacía
            if s == "":
                errs.append(
                    (
                        i,
                        c,
                        col_idx,
                        "mac_vacia",
                        "MAC es obligatoria y no puede estar vacía.",
                        "error",
                    )
                )
                continue

            # Debe cumplir XX:XX:XX:XX:XX:XX en hex mayúscula
            if not MAC_REGEX.fullmatch(s):
                errs.append(
                    (
                        i,
                        c,
                        col_idx,
                        "mac_invalida",
                        "MAC inválida: use solo hexadecimales en mayúscula y formato XX:XX:XX:XX:XX:XX (ej. 00:AF:2B:3C:4D:5E).",
                        "error",
                    )
                )

    # ================== VLAN OPCIONAL PERO ENTERA ==================
    if "VLAN" in df.columns:
        c = "VLAN"
        col_idx = df.columns.get_loc(c)
        for i, v in df[c].items():
            if pd.isna(v) or str(v).strip() == "":
                continue  # opcional

            val = str(v).strip()

            # "1003.0" lo damos como válido (se corrige después)
            if re.fullmatch(r"\d+\.0+", val):
                continue

            # Solo enteros puros
            if not re.fullmatch(r"\d+", val):
                errs.append(
                    (
                        i,
                        c,
                        col_idx,
                        "vlan_invalida",
                        "VLAN debe ser número entero (ej: 1003), sin decimales ni letras.",
                        "error",
                    )
                )

    # ================== NUMERO DE SERIE ONT ==================
    ONT_COL = "NUMERO DE SERIE ONT"
    if ONT_COL in df.columns:
        col_idx = df.columns.get_loc(ONT_COL)

        for i, v in df[ONT_COL].items():
            raw = "" if pd.isna(v) else str(v)

            cleaned = re.sub(r"[^A-Za-z0-9]", "", raw).upper()
            df.at[i, ONT_COL] = cleaned  # se guarda limpio

            if cleaned == "":
                errs.append(
                    (
                        i,
                        ONT_COL,
                        col_idx,
                        "ont_vacio",
                        "NUMERO DE SERIE ONT es obligatorio.",
                        "error",
                    )
                )
                continue

            if len(cleaned) not in (12, 16):
                errs.append(
                    (
                        i,
                        ONT_COL,
                        col_idx,
                        "ont_longitud_invalida",
                        f"Debe tener exactamente 12 o 16 caracteres (actual: {len(cleaned)}).",
                        "error",
                    )
                )
                continue

            if not re.fullmatch(r"[A-Za-z0-9]+", cleaned):
                errs.append(
                    (
                        i,
                        ONT_COL, 
                        col_idx,
                        "ont_formato_invalido",
                        "Debe contener solo letras y números.",
                        "error",
                    )
                )

    # ================== CAMPOS OBLIGATORIOS GENERALES ==================
    mandatory_cols = [
        "* CODIGO CONTRATO",
        "* NOMBRES",
        "* APELLIDOS",
        "* NOMBRES / RAZON SOCIAL",
        "* ESTADO",
        "* TIPO DOCUMENTO",
        "* N. DOCUMENTO",
        "* E-MAIL",
        "TELEFONOS",
        "* DURACION (MESES)",
        "* FECHA NACIMIENTO",
        "* FECHA CONTRATO",
        "* FECHA FIRMA",
        "FECHA ULTIMO CORTE",
        "DIA COBRO",
        "DIA CORTE",
        "PLAN DE INTERNET",
        "PRECIO DE INTERNET",
        "NUMERO DE SERIE ONT",
        "CALLE / DIRECCION",
        "IP",
        "VLAN",
        "MAC",
        "NAP",
        "OLT",
        "MIKROTIK",
        "DIRECCION",
    ]

    for col_name in mandatory_cols:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name)
            series = df[col_name]
            for i, v in series.items():
                s = "" if pd.isna(v) else str(v).strip()
                if s == "" or s.lower() in {"none", "nan", "null"}:
                    errs.append(
                        (
                            i,
                            col_name,
                            col_idx,
                            "campo_obligatorio",
                            f"{col_name} es obligatorio y no puede estar vacío.",
                            "error",
                        )
                    )

    if not errs:
        return pd.DataFrame(columns=["fila", "columna", "col_idx", "regla", "detalle", "severity"])


    return pd.DataFrame(
        errs,
        columns=["fila", "columna", "col_idx", "regla", "detalle", "severity"],
    )

# ---------- Estilos de errores ----------
def build_error_mask_by_position(df: pd.DataFrame, errores: pd.DataFrame):
    mask = pd.DataFrame("", index=df.index, columns=range(len(df.columns)))
    if errores is None or errores.empty:
        return mask
    for _, r in errores.iterrows():
        fila = int(r["fila"])
        # usa col_idx si existe; si no, marca todas las coincidencias por nombre
        if "col_idx" in r and not pd.isna(r["col_idx"]):
            j = int(r["col_idx"])
            if 0 <= j < len(df.columns) and fila in df.index:
                mask.iat[fila, j] = str(r.get("severity", "error"))
        else:
            # fallback por nombre
            col_name = str(r["columna"])
            for j, c in enumerate(df.columns):
                if str(c) == col_name and fila in df.index:
                    mask.iat[fila, j] = str(r.get("severity", "error"))
    return mask

def idx_to_excel_col(idx: int) -> str:
    """
    Convierte un índice 0-based (0=A, 1=B, ...) en letra de columna Excel (A, B, ..., AA, AB, ...).
    """
    if pd.isna(idx):
        return ""
    n = int(idx) + 1  # pasamos a 1-based
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result

# --- Helper para obtener coordenadas (fila, col_idx) de errores ---
def get_error_coords(errores: pd.DataFrame):
    """
    Devuelve un conjunto de coordenadas (fila, col_idx) de las celdas con error.
    Solo usa col_idx cuando está presente y no es NaN.
    """
    coords = set()
    if errores is None or errores.empty:
        return coords

    has_col_idx = "col_idx" in errores.columns

    for _, r in errores.iterrows():
        try:
            fila = int(r["fila"])
        except Exception:
            continue

        if has_col_idx and not pd.isna(r["col_idx"]):
            try:
                col = int(r["col_idx"])
            except Exception:
                continue
            coords.add((fila, col))

    return coords

def style_error_cells_ui(df: pd.DataFrame, errores: pd.DataFrame):
    # Creamos una copia con columnas únicas SOLO para la UI
    disp = ui_df_with_unique_columns(df)
    # Índice visible para el usuario: 1, 2, 3...
    disp.index = disp.index + 1

    # Máscara de errores sigue 0-based (coincide con el DF original)
    pos_mask = build_error_mask_by_position(df, errores)

    # Coordenadas revalidadas (celdas que ANTES tenían error y ahora ya no)
    resolved_coords = st.session_state.get("resolved_coords", set())

    def apply_colors(row):
        i_display = row.name      # 1, 2, 3, ...
        i0 = int(i_display) - 1   # índice real 0-based

        styles = []
        for j in range(len(row)):
            # severidad actual (error/warn/info) sobre esta vista
            if 0 <= i0 < pos_mask.shape[0]:
                sev = pos_mask.iat[i0, j]
            else:
                sev = ""

            if sev == "info":
                styles.append("background-color: #e7f7e7")
            elif sev == "warn":
                styles.append("background-color: #fff3cd")
            elif sev == "error":
                styles.append("background-color: #ffd6d6")
            else:
                # Si ya NO tiene error, pero su coordenada estuvo en errores previos -> marcado verde
                if (i0, j) in resolved_coords:
                    styles.append("background-color: #d4edda")  # verde suave (revalidado)
                else:
                    styles.append("")
        return styles

    return disp.style.apply(apply_colors, axis=1)


# ---------- Exportar Excel ----------
def _blank_unnamed_headers(ws, header_row_index: int, col_names):
    """Deja en blanco los encabezados que empiezan por 'Unnamed' en el archivo."""
    for i, name in enumerate(col_names, start=1):
        if str(name).lower().startswith("unnamed"):
            ws.cell(row=header_row_index, column=i).value = ""

def write_excel_with_errors(df, errores, rules_pkg, title_for_unnamed="INFORMACION DEL ABONADO"):
    """
    Exporta el MAESTRO con:
    - Limpieza de NA/NaN/<NA>/NaT -> None (celdas vacías en Excel)
    - Marcado de celdas con error (borde + fondo rojo claro)
    - Formato de fechas DD/MM/YYYY
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "MAESTRO"

    # 1) Aseguramos tipos datetime donde corresponde (para que Excel entienda bien las fechas)
    df_to_write = ensure_datetime_columns(df.copy())

    # 2) Reemplazar TODOS los valores nulos (NaN, <NA>, NaT, None) por None
    #    pd.notna(...) es False para NaN, NaT y <NA>, True para valores reales
    df_to_write = df_to_write.where(pd.notna(df_to_write), None)

    # 3) Volcar el DataFrame a la hoja, limpiando cualquier valor problemático en la fila
    for row in dataframe_to_rows(df_to_write, index=False, header=True):
        clean_row = []
        for v in row:
            # Cualquier cosa "nula" la convertimos a None (celda vacía)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                clean_row.append(None)
            else:
                clean_row.append(v)
        ws.append(clean_row)

    col_names = list(df_to_write.columns)

    # 4) Insertar fila de título si hay columnas "Unnamed"
    unnamed_idxs = [i for i, c in enumerate(col_names, start=1) if str(c).lower().startswith("unnamed")]
    if unnamed_idxs:
        ws.insert_rows(1)
        start = min(unnamed_idxs); end = max(unnamed_idxs)
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start)
        cell.value = title_for_unnamed
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        ws.row_dimensions[1].height = 22

    header_row_index = 2 if unnamed_idxs else 1
    _blank_unnamed_headers(ws, header_row_index, col_names)

    # 5) Estilos para marcar errores
    red = Side(style="thin", color="FF0000")
    red_border = Border(left=red, right=red, top=red, bottom=red)
    light_red_fill = PatternFill(fill_type="solid", fgColor="FFFFC7CE")  # rojo claro

    if errores is not None and not errores.empty:
        for _, r in errores.iterrows():
            try:
                fila1 = int(r["fila"])
            except Exception:
                continue

            # Usamos col_idx si está disponible
            if "col_idx" in r and not pd.isna(r["col_idx"]):
                try:
                    col_idx0 = int(r["col_idx"])
                except Exception:
                    continue

                if 0 <= col_idx0 < len(col_names):
                    excel_row = fila1 + header_row_index + 1
                    excel_col = col_idx0 + 1
                    cell = ws.cell(row=excel_row, column=excel_col)
                    cell.border = red_border
                    cell.fill = light_red_fill
            else:
                # Fallback: por nombre de columna
                col_name = str(r["columna"])
                matches = [j for j, name in enumerate(col_names) if str(name) == col_name]
                for col_idx0 in matches:
                    excel_row = fila1 + header_row_index + 1
                    excel_col = col_idx0 + 1
                    cell = ws.cell(row=excel_row, column=excel_col)
                    cell.border = red_border
                    cell.fill = light_red_fill

    # 6) Ajuste de ancho de columnas
    for i, c in enumerate(col_names, start=1):
        col_letter = ws.cell(row=header_row_index, column=i).column_letter
        ws.column_dimensions[col_letter].width = max(12, min(35, len(str(c)) + 4))

    # 7) Formato de fechas DD/MM/YYYY
    date_cols_by_name = date_columns_from_rules(rules_pkg)
    for pos in DATE_POSITIONS:
        if pos < len(col_names):
            date_cols_by_name.append(col_names[pos])

    if date_cols_by_name:
        name_to_idx = {name: idx for idx, name in enumerate(col_names, start=1)}
        for dc in date_cols_by_name:
            if dc in name_to_idx:
                cidx = name_to_idx[dc]
                for r in range(header_row_index + 1, ws.max_row + 1):
                    cell = ws.cell(row=r, column=cidx)
                    if cell.value is not None:
                        cell.number_format = "DD/MM/YYYY"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def write_excel_validated(df: pd.DataFrame, rules_pkg, title_for_unnamed="INFORMACION DEL ABONADO"):
    return write_excel_with_errors(df, errores=pd.DataFrame(), rules_pkg=rules_pkg, title_for_unnamed=title_for_unnamed)

# ---------- Helper para elegir DF vivo ----------
def pick_live_df():
    """Devuelve el DF que debe usarse: el editado si existe, sino el actual."""
    ed = st.session_state.get("edited_df", None)
    cur = st.session_state.get("current_df", None)
    return ed if ed is not None else cur

# ---------- Interfaz ----------
st.sidebar.subheader("Opciones")
enable_autocorrect = st.sidebar.checkbox("Autocorregir dominios de correo", value=True)

archivo = st.file_uploader("Sube el Excel (hoja 'MAESTRO')", type=["xlsx", "xls"])

if "current_df" not in st.session_state:
    st.session_state["current_df"] = None
if "edited_df" not in st.session_state:
    st.session_state["edited_df"] = None
if "file_id" not in st.session_state:
    st.session_state["file_id"] = None
if "resolved_coords" not in st.session_state:
    st.session_state["resolved_coords"] = set()
if "last_errores" not in st.session_state:
    st.session_state["last_errores"] = None
# Nuevos estados para mejorar rendimiento de validación
if "errores" not in st.session_state:
    st.session_state["errores"] = None
if "needs_validation" not in st.session_state:
    st.session_state["needs_validation"] = False
if "tipo_errs_all" not in st.session_state:
    st.session_state["tipo_errs_all"] = pd.DataFrame(columns=["fila","columna","col_idx","regla","detalle","severity"])
if "corrections_df" not in st.session_state:
    st.session_state["corrections_df"] = pd.DataFrame()
if "info_email_errors" not in st.session_state:
    st.session_state["info_email_errors"] = pd.DataFrame(columns=["fila","columna","col_idx","regla","detalle","severity"])



if archivo:
    try:
        # Identificador simple del archivo (nombre + tamaño)
        file_id = f"{archivo.name}_{archivo.size}"

        # Si es un archivo nuevo, reiniciamos; si es el mismo, conservamos DF vivo
        if st.session_state["file_id"] != file_id:
            st.session_state["file_id"] = file_id
            st.session_state["current_df"] = None
            st.session_state["edited_df"] = None
            st.session_state["resolved_coords"] = set()
            st.session_state["last_errores"] = None

        # --- LECTURA INTELIGENTE DESDE CERO SOLO SI AÚN NO HAY DF VIVO ---
        if st.session_state["current_df"] is None:
            df_raw = smart_read_maestro(archivo)

            # Normalización base
            df_norm  = apply_non_destructive_normalization(df_raw)
            df_fixed = apply_auto_fixes(df_norm, rules_pkg)

            # Mayúsculas primeras columnas + conteo
            df_fixed, count_upper = force_uppercase_first_cols(df_fixed, n=UPPERCASE_N)

            # Inferir tipo documento
            df_fixed, tipo_errs_all = infer_tipo_documento_from_docnum(
                df_fixed, TIPO_DOC_POS, DOC_NUM_POS
            )

            # Normalización de texto (acentos/ñ) + conteo
            df_fixed, count_norm = normalize_text_apostrophe(df_fixed)

            # Solo dígitos en N. DOCUMENTO
            df_fixed = digits_only_column(df_fixed, DOC_NUM_POS)

            if "MAC" in df_fixed.columns:
                df_fixed["MAC"] = df_fixed["MAC"].map(clean_mac)

            df_fixed = normalize_vlan(df_fixed)
            df_fixed, corrections_df, info_email_errors = apply_email_autocorrect(
                df_fixed, EMAIL_POS, enable_autocorrect
            )
            df_fixed = convert_int_columns(df_fixed, INT_COLUMNS)

            # Guardar DF vivo
            st.session_state["current_df"] = df_fixed.copy()
            st.session_state["edited_df"]  = None

            # Guardar info auxiliar en sesión
            st.session_state["tipo_errs_all"] = tipo_errs_all.copy()
            st.session_state["corrections_df"] = corrections_df.copy()
            st.session_state["info_email_errors"] = info_email_errors.copy()

            # 👉 Guardar contadores de correcciones automáticas
            st.session_state["count_uppercase"] = int(count_upper)
            st.session_state["count_text_normalized"] = int(count_norm)

            # También podemos guardar cuántos e-mails se corrigieron
            st.session_state["count_email_autocorrect"] = int(len(corrections_df))

            # Marcar que es necesario validar
            st.session_state["needs_validation"] = True

        else:
            # Si ya tenemos DF vivo, seguimos usando ese
            df_fixed = pick_live_df().copy()
            # Recuperamos info auxiliar de sesión (si existe)
            tipo_errs_all = st.session_state.get(
                "tipo_errs_all",
                pd.DataFrame(columns=["fila","columna","col_idx","regla","detalle","severity"])
            )
            corrections_df = st.session_state.get("corrections_df", pd.DataFrame())
            info_email_errors = st.session_state.get(
                "info_email_errors",
                pd.DataFrame(columns=["fila","columna","col_idx","regla","detalle","severity"])
            )

                # ---------------- VALIDACIÓN (solo cuando es necesario) ----------------
        if st.session_state.get("needs_validation", True):
            df_valid = pick_live_df().copy()
            df_valid.index = range(len(df_valid))

            errores_yaml = validate_dataframe(df_valid, rules_pkg)
            errores_pos  = validate_position_rules(df_valid)

            tipo_errs = tipo_errs_all.copy()
            if not tipo_errs.empty:
                tipo_errs["fila"] = tipo_errs["fila"].astype(int)

            frames = []
            for e in (errores_yaml, errores_pos, tipo_errs, info_email_errors):
                if e is not None and not e.empty:
                    if "severity" not in e.columns:
                        e = e.copy()
                        e["severity"] = "error"
                    frames.append(e)

            errores = pd.concat(
                frames,
                ignore_index=True
            ) if frames else pd.DataFrame(
                columns=["fila","columna","col_idx","regla","detalle","severity"]
            )
            # ✅ Deduplicar por celda para evitar 2 errores en la misma coordenada
            errores = dedupe_errors_df(errores)


            # Cachear resultado en sesión
            st.session_state["errores"] = errores.copy()
            st.session_state["needs_validation"] = False
        else:
            # Reusar errores ya calculados
            errores = st.session_state.get(
                "errores",
                pd.DataFrame(columns=["fila","columna","col_idx","regla","detalle","severity"])
            )


        # --------- Cálculo de celdas revalidadas (antes tenían error, ahora no) ---------
        prev_errores = st.session_state.get("last_errores", None)

        coords_actuales = get_error_coords(errores)
        coords_previos  = get_error_coords(prev_errores) if prev_errores is not None else set()

        nuevos_revalidados = coords_previos - coords_actuales

        prev_revalidados = st.session_state.get("resolved_coords", set())
        resolved_coords = prev_revalidados.union(nuevos_revalidados)

        st.session_state["resolved_coords"] = resolved_coords
        st.session_state["last_errores"] = errores.copy()

        # =================== SIGUE TODO LO DEMÁS ===================

        # --- VISTA PREVIA con sombreado ---
        st.subheader("Vista previa Informacion (🔴 Error, 🟡 Aviso, 🟢 Info, ✅ Revalidado en verde)")
        df_preview = convert_int_columns(
            format_dates_for_display(pick_live_df().head(50), rules_pkg),
            INT_COLUMNS
        )     
        st.dataframe(style_error_cells_ui(df_preview, errores), use_container_width=True)

        # Estructura (compara con nombres REALES pero normalizados para pequeños detalles)
        if COLUMN_ORDER is not None:
            expected_cols = list(COLUMN_ORDER)
            received_cols = list(st.session_state["current_df"].columns)

            ok_structure, diff_df = compare_columns_with_details(expected_cols, received_cols)

            if not ok_structure:
                st.error("❌ Las columnas NO coinciden exactamente con el formato esperado.")
                st.warning(f"Revisa la hoja **'{MAESTRO_SHEET}'** del archivo subido.")

                with st.expander("Ver detalle de columnas"):
                    # --- Tabla resumida de diferencias ---
                    if not diff_df.empty:
                        diff_view = diff_df.copy()

                        # Añadir letra de columna Excel
                        diff_view["columna_excel"] = diff_view["posición"].apply(idx_to_excel_col)

                        # Reordenar y eliminar columnas *_norm que no necesitas
                        diff_view = diff_view[["columna_excel", "posición", "esperado", "recibido"]]

                        st.write("Diferencias por posición (solo donde NO coinciden):")
                        st.dataframe(diff_view, use_container_width=True)
                        st.markdown("---")

                        # --- Listado completo, con letras y resaltando columnas problemáticas ---
                        st.write("Listado completo de columnas esperadas:")
                        st.markdown("<br>".join(highlight_column_list(expected_cols, diff_df)), unsafe_allow_html=True)

                        st.write("Listado completo de columnas recibidas:")
                        st.markdown("<br>".join(highlight_column_list(received_cols, diff_df)), unsafe_allow_html=True)

                        st.stop()

            else:
                st.success("✅ Estructura OK (columnas y orden coherentes).")

        # Correcciones de correo aplicadas (info)
        if enable_autocorrect and corrections_df is not None and not corrections_df.empty:
            corr_show = corrections_df.copy()
            corr_show["fila"] = corr_show["fila"].astype(int)

            # Recuperar contadores de sesión
            count_upper = int(st.session_state.get("count_uppercase", 0) or 0)
            count_norm  = int(st.session_state.get("count_text_normalized", 0) or 0)
            count_email = int(st.session_state.get("count_email_autocorrect", len(corr_show)) or 0)

            # Total de correcciones automáticas
            total_auto = count_upper + count_norm + count_email

            # Texto con salto de línea
            resumen = (
                f"Correcciones automáticas totales: {total_auto} celdas modificadas.\n"
                f"MAYÚSCULAS: {count_upper}, sin acentos/ñ: {count_norm}, "
                f"correcciones de e-mail: {count_email}."
            )

            st.warning(resumen)
            st.dataframe(corr_show, use_container_width=True)

        # ✅ Blindaje: usa siempre el errores final de sesión (dedupe + cache)
        errores = st.session_state.get(
            "errores",
            errores if errores is not None else pd.DataFrame(columns=["fila","columna","col_idx","regla","detalle","severity"])
        )

        # Normaliza si por algún motivo viene None
        if errores is None:
            errores = pd.DataFrame(columns=["fila","columna","col_idx","regla","detalle","severity"])


        # ---- Errores ----
              
        if not errores.empty:
            # Indicar la hoja donde se están encontrando los errores
            st.subheader("Vista Previa Incidencias")
            st.info(f"Incidencias detectadas en la hoja de Excel: **'{MAESTRO_SHEET}'**")

            # Copia de trabajo
            errores_mostrar = errores.copy()

            # Rellenar col_idx faltantes usando el nombre de columna
            if "col_idx" in errores_mostrar.columns:
                # Mapa nombre -> índice real del DF vivo
                name_to_idx = {
                    str(c): j for j, c in enumerate(st.session_state["current_df"].columns)
                }

                # Solo filas donde col_idx viene vacío / NaN
                mask_na = errores_mostrar["col_idx"].isna()
                if mask_na.any():
                    errores_mostrar.loc[mask_na, "col_idx"] = (
                        errores_mostrar.loc[mask_na, "columna"]
                        .astype(str)
                        .map(name_to_idx)
                    )

                # Aseguramos tipo entero nullable y calculamos letra de columna Excel
                errores_mostrar["col_idx"] = errores_mostrar["col_idx"].astype("Int64")
                errores_mostrar["columna_excel"] = errores_mostrar["col_idx"].apply(
                    idx_to_excel_col
                )

            # Severidad en español
            errores_mostrar["Importancia"] = errores_mostrar["severity"].str.lower().map(
                SEVERITY_MAP_ES
            )

            # Reordenar para que sea coordenada fácil: Columna (Excel) + Fila
            cols = list(errores_mostrar.columns)
            new_order = []
            if "columna_excel" in cols:
                new_order.append("columna_excel")  # primero la columna Excel
            if "fila" in cols:
                new_order.append("fila")  # luego la fila → coordenada tipo G | 15
            # el resto
            new_order += [c for c in cols if c not in new_order]
            errores_mostrar = errores_mostrar[new_order]

            # Índice visible 1,2,3...
            errores_mostrar.index = errores_mostrar.index + 1

            # Ocultar columnas internas en la vista (pero NO borrarlas del DF original)
            cols_visibles = [
                c
                for c in errores_mostrar.columns
                if c not in ("col_idx", "idxinterno", "severity")
            ]

            st.warning(f"⚠️ Se encontraron {len(errores)} incidencias.")
            st.dataframe(errores_mostrar[cols_visibles], use_container_width=True)

            # --- Botón extra: descargar MAESTRO con SOLO las celdas con incidencias marcadas en rojo ---
            st.info(
                "También puedes descargar un Excel con todas las filas pero "
                "solo las celdas con incidencias sombreadas en rojo."
            )
            if st.button("🔴 Generar Excel con celdas en error", key="btn_generate_errors_file"):
                df_export_err = pick_live_df().copy()
                xlsx_bytes_err = write_excel_with_errors(
                    df_export_err, errores, rules_pkg, title_for_unnamed="INFORMACION DEL ABONADO"
                )
                st.session_state["xlsx_errors_file"] = xlsx_bytes_err

            # Si el archivo ya está generado → mostrar botón para descargarlo
            if "xlsx_errors_file" in st.session_state:
                st.download_button(
                    "📁 Descargar MAESTRO_errores.xlsx",
                    st.session_state["xlsx_errors_file"],
                    "MAESTRO_errores.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_download_errors_file",
                )


           

            # ============== EDITOR DE INCIDENCIAS (CELDA POR CELDA) ==============
            st.markdown("## Editor de incidencias (por fila/columna)")

            # Solo filas con col_idx válido
            if "col_idx" in errores.columns:
                errores_editables = errores.dropna(subset=["col_idx"]).copy()
            else:
                errores_editables = pd.DataFrame(columns=errores.columns)

            if errores_editables.empty:
                st.info("No hay celdas editables con posición de columna definida.")
            else:
                errores_editables["fila"] = errores_editables["fila"].astype(int)
                errores_editables["col_idx"] = errores_editables["col_idx"].astype(int)

                # Convertir a letra de Excel
                errores_editables["columna_excel"] = errores_editables["col_idx"].apply(idx_to_excel_col)
                errores_editables["fila_excel"] = errores_editables["fila"] + 1

                def severity_icon(sev: str) -> str:
                    sev = str(sev).lower()
                    if sev == "error":
                        return "🔴 ERROR"
                    if sev == "warn":
                        return "🟡 AVISO"
                    if sev == "info":
                        return "🟢 INFO"
                    return "⚪"

                errores_editables["estado"] = errores_editables["severity"].map(severity_icon)

                # Valores actuales desde DF vivo
                valores_actuales = []
                for _, r in errores_editables.iterrows():
                    fila_i = int(r["fila"])
                    col_i = int(r["col_idx"])
                    val = st.session_state["current_df"].iat[fila_i, col_i]
                    valores_actuales.append(val)

                errores_editables["valor_actual"] = valores_actuales

                # Forzar a texto para que sea compatible con st.column_config.TextColumn
                errores_editables["valor_actual"] = errores_editables["valor_actual"].apply(
                    lambda x: "" if pd.isna(x) else str(x)
                ).astype("string")

                errores_editables["nuevo_valor"] = errores_editables["valor_actual"].copy()

                # Columnas que muestra el editor
                editor_cols = [
                    "estado",
                    "columna_excel",
                    "fila_excel",   # 👈 visible al usuario
                    "fila",         # 👈 interna (0-based)
                    "columna",
                    "col_idx",      # 👈 interna
                    "valor_actual",
                    "detalle",
                    "severity",
                    "nuevo_valor",
                ]

                editor_view = errores_editables[editor_cols].copy()
                #  índice visible 1,2,3... en lugar de 0,1,2...
                editor_view.index = editor_view.index + 1
                st.session_state["editor_base"] = editor_view.copy()

                # Solo estas columnas se muestran en la UI
                visible_cols = [
                    "estado",
                    "columna_excel",
                    "fila_excel",   # Fila 1,2,3...
                    "columna",
                    "valor_actual",
                    "detalle",
                    "severity",
                    "nuevo_valor",
                ]

                edited_issues = st.data_editor(
                    editor_view,
                    use_container_width=True,
                    num_rows="fixed",
                    key="editor_issues",
                    column_order=visible_cols,   # 👈 aquí ocultamos fila / col_idx internas
                    column_config={
                        "estado": st.column_config.TextColumn("Estado", disabled=True),
                        "columna_excel": st.column_config.TextColumn("Columna (Excel)", disabled=True),
                        "fila_excel": st.column_config.NumberColumn("Fila", disabled=True),
                        "columna": st.column_config.TextColumn("Nombre columna", disabled=True),
                        "valor_actual": st.column_config.TextColumn("Valor actual", disabled=True),
                        "detalle": st.column_config.TextColumn("Detalle", disabled=True),
                        "severity": st.column_config.TextColumn("Severidad", disabled=True),
                        "nuevo_valor": st.column_config.TextColumn("Nuevo valor", disabled=False),
                    },
                )

                # === BLINDAJE: restaurar todas las columnas excepto 'nuevo_valor' ===
                base_view = st.session_state.get("editor_base", editor_view)
                safe_edited = edited_issues.copy()

                for col in editor_cols:
                    if col != "nuevo_valor":
                        # Cualquier cambio manual en estas columnas se ignora
                        safe_edited[col] = base_view[col]

                # Aplicar cambios solo de 'nuevo_valor' al DF vivo
                df_editado = st.session_state["current_df"].copy()

                for idx, row in safe_edited.iterrows():
                    fila_i = int(base_view.loc[idx, "fila"])
                    col_i = int(base_view.loc[idx, "col_idx"])
                    new_val = row["nuevo_valor"]

                    # Si es la columna MAC, normalizamos (hex + mayúscula + ':')
                    col_name = df_editado.columns[col_i]
                    if str(col_name).strip().upper() == "MAC":
                        new_val = clean_mac(new_val)

                    if 0 <= fila_i < df_editado.shape[0] and 0 <= col_i < df_editado.shape[1]:
                        df_editado.iat[fila_i, col_i] = new_val
                st.session_state["needs_validation"] = True
                st.session_state["current_df"] = df_editado.copy()
                st.session_state["edited_df"] = df_editado.copy()

            # Descarga "de todos modos" ANTES de revalidar
            st.info("¿Deseas descargar el MAESTRO corregido aunque aún haya incidencias?")
            confirm_anytime = st.checkbox("Sí, descargar MAESTRO_corregido.xlsx con posibles errores.", key="dl_anytime")
            if confirm_anytime:
                df_export_any = pick_live_df().copy()
                if enable_autocorrect:
                    df_export_any, _, _ = apply_email_autocorrect(df_export_any, EMAIL_POS, True)
                st.download_button(
                    "⬇️ Descargar MAESTRO_corregido.xlsx (con posibles errores)",
                    write_excel_validated(df_export_any, rules_pkg, title_for_unnamed="INFORMACION DEL ABONADO"),
                    "MAESTRO_corregido.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_anytime_btn"
                )

            # Revalidar (respeta edición)
            st.markdown("---")
        if st.button("Revalidar"):
            # ✅ Partimos del DF vivo (incluye lo editado)
            df2 = pick_live_df().copy()

            # Forzar PAIS = VENEZUELA (si existe la columna)
            for country_col_name in ["PAIS", "* PAIS"]:
                if country_col_name in df2.columns:
                    df2[country_col_name] = "VENEZUELA"

            # ✅ Aplicar pipeline completo y consistente
            df2, tipo_errs_all_new, corrections_df_new, info_email_errors_new = run_full_pipeline(df2, enable_autocorrect)

            # Guardar DF vivo
            st.session_state["current_df"] = df2.copy()
            st.session_state["edited_df"] = None

            # ✅ Refrescar auxiliares para que no queden errores “viejos”
            st.session_state["tipo_errs_all"] = tipo_errs_all_new.copy()
            st.session_state["corrections_df"] = corrections_df_new.copy()
            st.session_state["info_email_errors"] = info_email_errors_new.copy()

            # ✅ Recalcular errores en el próximo ciclo
            st.session_state["needs_validation"] = True
            st.session_state["errores"] = None

            st.rerun()


        

    except Exception as e:
        st.exception(e)
else:
    st.info("Sube un archivo para validar.")

# ====================== FOOTER HORIZONTAL CON LOGO FOXBYTE ======================

import base64
from pathlib import Path

# Detectar carpeta /assets
BASE_DIR = Path(__file__).parent
assets_dir = BASE_DIR / "assets"

# Convertir logo Foxbyte a base64
footer_logo_b64 = None
if assets_dir.exists():
    candidates = list(assets_dir.glob("logo_foxbyte*"))
    if candidates:
        with open(candidates[0], "rb") as f:
            footer_logo_b64 = base64.b64encode(f.read()).decode()

# CSS + HTML del footer
footer_html = f"""
<style>
.app-footer {{
    position: fixed;
    bottom: 0;
    left: 0;
    width: 100%;

    background-color: rgba(255,255,255,0.95);
    border-top: 1px solid #ddd;

    display: flex;
    align-items: center;
    justify-content: center;   /* Centrado perfecto */

    padding: 6px 0;
    z-index: 999;
}}

.footer-inner {{
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 10px;   /* espacio entre logo y texto */
}}

.footer-img {{
    height: 50px;   /* tamaño del logo (ajustable) */
    opacity: 0.9;
}}

.footer-text {{
    font-size: 18px;
    color: #666;
}}
</style>

<div class="app-footer">
    <div class="footer-inner">
        {f'<img src="data:image/png;base64,{footer_logo_b64}" class="footer-img">' if footer_logo_b64 else ""}
        <div class="footer-text">
            Desarrollado por <strong>Foxbyte</strong>  <strong>2025</strong>
        </div>
    </div>
</div>
"""

st.markdown(footer_html, unsafe_allow_html=True)

# Ocultar ícono de GitHub de la barra superior de Streamlit
st.markdown(
    """
    <style>
    /* Oculta el botón de GitHub en la barra superior */
    button[kind="header"] svg[data-testid="stIconGithub"] {
        display: none !important;
    }
    /* Opcional: también oculta el texto si apareciera */
    a[href*="github.com"] {
        text-decoration: none;
    }
    </style>
    """,
    unsafe_allow_html=True
)
